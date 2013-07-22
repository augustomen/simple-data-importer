# coding: utf-8
from openpyxl.reader.excel import load_workbook

from simpledataimporter.utils import slugify


class XLSXReader(object):
    """ Iterable reader for use with BaseImporter.
        Uses openpyxl iterable reader instead of xlrd.
        Works only with Excel 2007- files (*.xlsx).

        filename can be a string or a file object.
        To change the current worksheet, use:
            xlsxreader.activate_worksheet(1)
        or
            xlsxreader.activate_worksheet('Sheet2')
        Field names are taken from the first row.
    """

    def __init__(self, filename):
        self.filename = filename
        self._workbook = load_workbook(filename, use_iterators=True)
        self._active_worksheet = None
        self._fieldnames = None

    def __len__(self):
        return self.worksheet.get_highest_row() - 1

    def __iter__(self):
        fields = self.fieldnames
        for i, row in enumerate(self.worksheet.iter_rows()):
            if i == 0:
                continue
            yield dict(zip(fields, [cell.internal_value for cell in row]))

    @property
    def fieldnames(self):
        if self._fieldnames is None:
            row = self.worksheet.iter_rows().next()
            self._fieldnames = []
            for cell in row:
                self._fieldnames.append(cell.internal_value)
        return self._fieldnames

    @property
    def workbook(self):
        return self._workbook

    def activate_worksheet(self, name_or_index):
        if isinstance(name_or_index, int):
            name_or_index = self._workbook.get_sheet_names()[name_or_index]
        self._active_worksheet = self.workbook.get_sheet_by_name(name_or_index)
        return self._active_worksheet

    @property
    def worksheet(self):
        if self._active_worksheet is None:
            self._active_worksheet = self._workbook.get_active_sheet()
        return self._active_worksheet


class SlugXLSXReader(XLSXReader):
    """ Same as XLSXReader, but all field names are slugified. """

    @property
    def fieldnames(self):
        if self._fieldnames is None:
            self._fieldnames = map(
                slugify, super(SlugXLSXReader, self).fieldnames)
        return self._fieldnames
